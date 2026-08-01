"""
================================================================================
檔案名稱: ui/pages/form_management/shared.py
功能說明: Form Management UI 共用 Helper、Schema 字典與渲染器
================================================================================
"""

import os
import json
import time
import math
import uuid
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from ui.pages.shared import build_admin_headers, resolve_api_base_url

_FORM_MGMT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # ui/pages
_PROJECT_ROOT = os.path.dirname(os.path.dirname(_FORM_MGMT_DIR))  # 專案根目錄
_LEGACY_TEMPLATES_DIR = os.path.join(_FORM_MGMT_DIR, "db", "templates")
_PRIMARY_TEMPLATES_DIR = os.path.join(_PROJECT_ROOT, "db", "templates")

JSON_TPL_PATH = os.path.join(_PROJECT_ROOT, "db", "form_templates.json")
TEMPLATES_DIR = _PRIMARY_TEMPLATES_DIR if os.path.isdir(_PRIMARY_TEMPLATES_DIR) else _LEGACY_TEMPLATES_DIR
CONTRACTS_DIR = os.path.join(_PRIMARY_TEMPLATES_DIR, "contracts") if os.path.isdir(_PRIMARY_TEMPLATES_DIR) else os.path.join(_LEGACY_TEMPLATES_DIR, "contracts")


def fetch_staff_contract_context(case_no: str, assignment_id: int | None = None) -> dict:
    """Read staff-contract facts from FastAPI without writing the workbook."""
    query = urlencode({"assignment_id": assignment_id}) if assignment_id else ""
    base_url = resolve_api_base_url()
    url = f"{base_url}/api/v1/contracts/staff/{case_no}"
    if query:
        url = f"{url}?{query}"
    request = Request(url=url, headers=build_admin_headers())
    try:
        with urlopen(request, timeout=5) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise ValueError(f"契約資料 API 回應 {error.code}: {detail}") from error
    except URLError as error:
        raise ValueError(f"無法連線契約資料 API ({base_url}): {error.reason}") from error

    if not payload.get("success", False) or not isinstance(payload.get("data"), dict):
        raise ValueError(payload.get("error") or payload.get("message") or "契約資料 API 未回傳資料")
    return payload["data"]


def flatten_staff_contract_context(context: dict) -> dict:
    """Flatten the read-only API payload into existing Excel mapping keys."""
    order = context.get("order") or {}
    client = context.get("client") or {}
    assignment = context.get("assignment") or {}
    staff = context.get("staff") or {}

    flat = {
        **{key: value for key, value in order.items() if value is not None},
        **{key: value for key, value in assignment.items() if value is not None},
        "client_name": client.get("name"),
        "client_phone": client.get("phone"),
        "city": client.get("city"),
        "address": client.get("address"),
        "service_type": client.get("service_type"),
        "service_time": client.get("service_time"),
        "staff_name": staff.get("name"),
        "staff_phone": staff.get("phone"),
    }
    return {key: value for key, value in flat.items() if value is not None}


def generate_field_id() -> str:
    """產生全域絕對不重複的 field_id (INV-UI-FORM-07 防碰撞防線)"""
    return f"f_{uuid.uuid4().hex[:10]}_{int(time.time() * 1000)}"


def get_html_hex_color(color_obj, default=None):
    """將 openpyxl Color 物件轉為標準 HTML HEX 色碼 (#RRGGBB)"""
    if not color_obj:
        return default
    rgb = getattr(color_obj, 'rgb', None)
    if not rgb:
        return default
    rgb_str = str(rgb).upper()
    if len(rgb_str) == 8: # AARRGGBB
        hex_val = rgb_str[2:]
        if hex_val == "000000" or hex_val == "FFFFFF":
            return default
        return f"#{hex_val}"
    elif len(rgb_str) == 6:
        return f"#{rgb_str}"
    return default


def get_border_style(cell_border):
    """精確解析 Excel 單元格顯式設定之邊框，空白單元格不繪製多餘灰框，確保列印 PDF 時 100% 保留實心框線 (INV-UI-FORM-29)"""
    if not cell_border:
        return "border: none !important;"

    b_left = "1px solid #111111 !important;" if (cell_border.left and cell_border.left.style) else "none !important;"
    b_right = "1px solid #111111 !important;" if (cell_border.right and cell_border.right.style) else "none !important;"
    b_top = "1px solid #111111 !important;" if (cell_border.top and cell_border.top.style) else "none !important;"
    b_bottom = "1px solid #111111 !important;" if (cell_border.bottom and cell_border.bottom.style) else "none !important;"

    return f"border-left: {b_left} border-right: {b_right} border-top: {b_top} border-bottom: {b_bottom}"


def render_excel_contract_mirror(contract_config: dict, target_order: dict, global_stats: dict) -> str:
    """實時 1:1 解析與高精度鏡像渲染 Excel 實體範本檔 (含背景色、欄寬、粗體與原廠邊框)"""
    tpl_filename = contract_config.get('template_filename', 'contract_client_copy.xlsx')
    excel_file_path = os.path.join(CONTRACTS_DIR, tpl_filename)

    if not os.path.exists(excel_file_path):
        return f"<div style='color:red;'>❌ 未找到 Excel 範本檔案: {tpl_filename}</div>"

    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active

    mappings = contract_config.get('param_mappings', {})

    max_cols = min(ws.max_column, 10)
    col_widths = {}
    total_w = 0
    for c in range(1, max_cols + 1):
        col_letter = get_column_letter(c)
        w = ws.column_dimensions[col_letter].width
        w_val = float(w) if w else 12.0
        col_widths[c] = w_val
        total_w += w_val

    colgroup_html = "<colgroup>\n"
    for c in range(1, max_cols + 1):
        pct = round((col_widths[c] / total_w) * 100, 2)
        colgroup_html += f"  <col style='width: {pct}%;'>\n"
    colgroup_html += "</colgroup>\n"

    merged_spans = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.min_col, rng.min_row, rng.max_col, rng.max_row
        top_left_coord = get_column_letter(min_col) + str(min_row)
        colspan = max_col - min_col + 1
        rowspan = max_row - min_row + 1

        merged_spans[top_left_coord] = (colspan, rowspan)

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell_coord = get_column_letter(c) + str(r)
                if cell_coord != top_left_coord:
                    merged_spans[cell_coord] = None

    table_rows_html = ""
    for r in range(1, min(ws.max_row + 1, 190)):
        row_html = ""
        for c in range(1, max_cols + 1):
            coord = get_column_letter(c) + str(r)

            if coord in merged_spans and merged_spans[coord] is None:
                continue

            cell = ws.cell(row=r, column=c)
            cell_val = str(cell.value).strip() if cell.value is not None else ""

            span_attr = ""
            if coord in merged_spans and merged_spans[coord] is not None:
                c_span, r_span = merged_spans[coord]
                if c_span > 1: span_attr += f" colspan='{c_span}'"
                if r_span > 1: span_attr += f" rowspan='{r_span}'"

            bg_color = get_html_hex_color(cell.fill.start_color if cell.fill else None, default="transparent")
            text_color = get_html_hex_color(cell.font.color if cell.font else None, default="#111111")
            is_bold = cell.font.bold if cell.font else False
            font_weight_css = "bold" if is_bold else "normal"
            border_css = get_border_style(cell.border)

            h_align = cell.alignment.horizontal if (cell.alignment and cell.alignment.horizontal) else "left"
            if h_align == "general": h_align = "left"

            is_excel_wrap = bool(cell.alignment and cell.alignment.wrap_text)
            wrap_css = "white-space:normal !important; word-break:break-word !important;" if is_excel_wrap else "white-space:nowrap; overflow:visible;"

            if coord in mappings:
                p_info = mappings[coord]
                db_k = p_info.get('db_key', '')

                bg_color = "transparent" if target_order else (bg_color if bg_color != "transparent" else "#FFFF00")
                text_color = "#0D47A1" if not target_order else "#111111"
                font_weight_css = "bold"

                if p_info.get('status') == 'pending':
                    cell_val = "暫不連動"
                    bg_color = "#FFF3CD"
                    text_color = "#9C6500"
                elif db_k == '__today__':
                    cell_val = date.today().strftime('%Y-%m-%d')
                elif db_k in global_stats:
                    cell_val = format_db_value(db_k, global_stats[db_k])
                elif target_order and db_k in target_order:
                    cell_val = format_db_value(db_k, target_order.get(db_k, ''))
                elif not cell_val or cell_val == "None":
                    cell_val = f"[{p_info.get('label', coord)}]"

            style_str = f"background:{bg_color}; color:{text_color}; font-weight:{font_weight_css}; text-align:{h_align}; vertical-align:middle; padding:3px 6px; {border_css} font-size:12px; {wrap_css}"
            row_html += f"<td {span_attr} style='{style_str}'>{cell_val}</td>"

        table_rows_html += f"<tr>{row_html}</tr>"

    mirror_html = f"""
    <div id="excel-mirror-container" style="font-family:'Segoe UI', Microsoft JhengHei, sans-serif; padding:15px; background:#FAFAFA; border-radius:8px;">
        <style>
            @page {{
                size: A4 portrait;
                margin: 8mm 10mm;
            }}
            @media print {{
                .no-print {{ display: none !important; }}
                #excel-mirror-container {{ padding: 0 !important; background: white !important; border-radius: 0 !important; margin: 0 !important; }}
                body, html {{ background: white !important; margin: 0 !important; padding: 0 !important; }}
                table, td, th {{
                    -webkit-print-color-adjust: exact !important;
                    print-color-adjust: exact !important;
                    color-adjust: exact !important;
                }}
            }}
        </style>
        <div style="text-align:center; margin-bottom:10px; color:#1565C0; font-weight:bold; font-size:15px;" class="no-print">
            📊 《{contract_config.get('name')}》 1:1 原汁原味 Excel 樣式鏡像渲染視窗 (檔案: {tpl_filename})
            <div style="font-size:12px; color:#666; font-weight:normal; margin-top:4px;">
                💡 提示：點擊下方列印按鈕，另存 PDF 時可在預覽視窗右側取消勾選【頁首和頁尾】即可阻斷頂部日期與 localhost 網址！
            </div>
        </div>
        <table style="border-collapse:collapse; width:100%; table-layout:fixed; background:#FFFFFF; margin:auto;">
            {colgroup_html}
            <tbody>
                {table_rows_html}
            </tbody>
        </table>
        <div style="text-align:center; margin-top:15px;" class="no-print">
            <button onclick="window.print()" style="background-color:#2E7D32; color:white; border:none; padding:8px 20px; font-size:14px; font-weight:bold; border-radius:6px; cursor:pointer;">
                🖨️ 點擊另存為 PDF / 輸出列印 (Print / Save as PDF)
            </button>
        </div>
    </div>
    """
    return mirror_html


def load_contract_templates():
    """從 db/templates/contracts/ 讀取所有制式契約與變數代理設定 (INV-UI-FORM-16/17)"""
    os.makedirs(CONTRACTS_DIR, exist_ok=True)
    c_list = []
    json_files = [f for f in os.listdir(CONTRACTS_DIR) if f.endswith('.json')]
    for fname in sorted(json_files):
        fpath = os.path.join(CONTRACTS_DIR, fname)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                c_list.append(json.load(f))
        except Exception:
            pass
    return c_list


def save_contract_template(contract_data: dict):
    """持久化保存契約變數對照設定至 db/templates/contracts/*.json"""
    os.makedirs(CONTRACTS_DIR, exist_ok=True)
    cid = contract_data.get('id', 'contract_hsinchu_v1')
    fpath = os.path.join(CONTRACTS_DIR, f"{cid}.json")
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(contract_data, f, ensure_ascii=False, indent=2)


def load_json_templates():
    """從 db/templates/ 目錄讀取所有獨立模板 JSON 檔案 (INV-UI-FORM-13 獨立模組架構 + 去重防護)"""
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    tpls = []

    json_files = [f for f in os.listdir(TEMPLATES_DIR) if f.endswith('.json')]
    if json_files:
        for fname in sorted(json_files):
            fpath = os.path.join(TEMPLATES_DIR, fname)
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    t = json.load(f)
                    seen_ids = set()
                    for idx, field in enumerate(t.get('fields', [])):
                        fid = field.get('id')
                        if not fid or fid in seen_ids:
                            field['id'] = generate_field_id()
                        seen_ids.add(field['id'])
                    tpls.append(t)
            except Exception:
                pass
        return tpls

    if os.path.exists(JSON_TPL_PATH):
        try:
            with open(JSON_TPL_PATH, "r", encoding="utf-8") as f:
                old_tpls = json.load(f)
                for t in old_tpls:
                    save_single_template(t)
                return old_tpls
        except Exception:
            pass

    return []


def save_single_template(template: dict):
    """將單一模板寫入獨立的 db/templates/tpl_xx.json 檔案 (INV-UI-FORM-13 按需獨立寫入)"""
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    tpl_id = template.get('id') or f"tpl_{int(time.time())}"
    template['id'] = tpl_id
    fpath = os.path.join(TEMPLATES_DIR, f"{tpl_id}.json")
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(template, f, ensure_ascii=False, indent=2)


def delete_single_template(tpl_id: str):
    """單獨刪除指定的 db/templates/tpl_xx.json 檔案 (INV-UI-FORM-13 獨立檔案刪除)"""
    fpath = os.path.join(TEMPLATES_DIR, f"{tpl_id}.json")
    if os.path.exists(fpath):
        try:
            os.remove(fpath)
        except Exception:
            pass


def save_json_templates(templates):
    """批次寫入所有獨立模板檔案 (向下相容)"""
    for t in templates:
        save_single_template(t)


def safe_int(val) -> int:
    if val is None:
        return 0
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return 0
        return int(round(f))
    except:
        return 0


def format_db_value(db_k: str, val_raw) -> str:
    """精確格式化 DB 36 欄位與 15 大照護細節數值 (INV-UI-FORM-12 精確單詞防護 Guardrail)"""
    if val_raw is None or str(val_raw).strip() == "":
        return "—"

    val_str_clean = str(val_raw).strip()
    db_k_lower = str(db_k).lower()

    if "date" in db_k_lower or "due" in db_k_lower:
        if isinstance(val_raw, (date, datetime)):
            return val_raw.strftime("%Y-%m-%d")
        return val_str_clean.split(" ")[0].strip()

    money_words = {'amount', 'fee', 'payable', 'rate', 'salary', 'deposit'}
    key_tokens = set(db_k_lower.split('_'))

    is_money_field = bool(key_tokens.intersection(money_words)) and not ('days' in db_k_lower or 'hours' in db_k_lower)

    if is_money_field:
        try:
            f = float(val_raw)
            if not (math.isnan(f) or math.isinf(f)):
                return f"{safe_int(f):,} 元"
        except (ValueError, TypeError):
            pass

    return val_str_clean


DB_TABLE_FIELDS = {
    "orders (訂單主表 - 36 大業務與金額 calculations)": {
        "case_no": "案件編號 (case_no)",
        "start_date": "預期服務開始日 (start_date)",
        "actual_start_date": "服務開始 (actual_start_date)",
        "actual_end_date": "服務結束 (actual_end_date)",
        "service_days": "希望服務天數 (service_days)",
        "service_hours_per_day": "每日服務時數 (service_hours_per_day)",
        "service_time": "服務時段 (service_time)",
        "service_mode": "服務方式 (service_mode)",
        "total_hours": "服務總時數 (total_hours)",
        "subsidy_hours": "補助時數 (subsidy_hours)",
        "self_pay_hours": "自費時數 (self_pay_hours)",
        "claim_total_days": "請款總日數 (claim_total_days)",
        "floor_fee": "樓層費用 (floor_fee)",
        "employer_hourly_rate": "雇主單價 (employer_hourly_rate)",
        "deposit_days": "訂金天數 (deposit_days)",
        "deposit_amount": "訂金金額 (deposit_amount)",
        "deposit_date": "訂金入帳日 (deposit_date)",
        "first_payment_days": "第一期款天數 (first_payment_days)",
        "first_payment_amount": "第一期金額 (first_payment_amount)",
        "first_payment_date": "第一期款入帳日 (first_payment_date)",
        "second_payment_days": "第二期款天數 (second_payment_days)",
        "second_payment_amount": "第二期金額 (second_payment_amount)",
        "second_payment_date": "第二期款入帳日 (second_payment_date)",
        "final_payment_date": "尾款日期 (final_payment_date)",
        "total_employer_self_pay_payable": "雇主自費合計金額 (total_employer_self_pay_payable)",
        "order_status": "訂單成立狀態 (order_status)",
        "staff_name": "服務人員 (staff_name)",
        "caregiver_rate": "服務單價 (caregiver_rate)",
        "special_holidays": "特殊休假 (special_holidays)",
        "service_salary": "服務薪資 (service_salary)",
        "salary_payment_date_1": "預計發薪日 (salary_payment_date_1)",
        "subsidy_salary": "補助薪資金額 (subsidy_salary)",
        "total_caregiver_salary": "總薪資 (total_caregiver_salary)",
        "govt_claim_date": "市府請款日 (govt_claim_date)"
    },
    "clients (客戶主表 - 個人基本資料與市府申請表)": {
        "client_name": "客戶名稱 (client_name)",
        "phone": "聯絡電話 (phone)",
        "address": "通訊與服務地址 (address)",
        "due_date": "預產期 (due_date)",
        "email": "Email 電子郵件 (email)",
        "bank_code": "雇主退款銀行代號/分行 (bank_code)",
        "bank_account": "雇主退款銀行帳號 (bank_account)",
        "service_mode": "服務方式 (service_mode)",
        "baby_info": "寶寶資訊 (baby_info)",
        "delivery_mode": "生產方式 (delivery_mode)",
        "residence_type": "居住型態 (residence_type)",
        "notes": "其他備註事項 (notes)",
        "id_card": "身分證字號 (id_card)",
        "line_id": "LINE ID (line_id)",
        "identity_status": "身分資格（唯讀） (identity_status)"
    },
    "staff (服務人員 - 月嫂主表與撥款帳號)": {
        "staff_name": "月嫂姓名 (staff_name)",
        "staff_phone": "月嫂電話 (staff_phone)",
        "staff_bank_code": "月嫂銀行代號/分行 (staff_bank_code)",
        "staff_bank_account": "月嫂銀行帳號 (staff_bank_account)",
        "id_card": "月嫂身分證號 (id_card)",
        "transportation": "服務交通工具 (transportation)",
        "service_region": "可承接案件區域 (service_region)"
    },
    "beclass_records (BeClass 報名表 - 15 大產婦照顧與飲食環境細節)": {
        "dietary_habits": "月子餐點調理喜好/飲食習慣",
        "vegetarian_preference": "無葷食時是否可接受蛋奶素餐食",
        "alcohol_ratio": "餐飲含酒比例 (全酒/半酒/無酒)",
        "cooking_oil_type": "料理用油 (苦茶油/麻油/橄欖油)",
        "maternal_allergy": "媽咪有無過敏體質",
        "special_care_notes": "特殊照護時應注意事項",
        "meal_preferences": "餐點喜忌備註",
        "cooking_tools": "現有烹煮工具 (炒菜鍋/電鍋/烤箱)",
        "bath_water_prep": "洗澡水準備方式 (熱水/薑水/中藥包)",
        "breastfeeding_method": "哺乳方式 (母乳/配方奶)",
        "holiday_pricing_terms": "特殊計費: 國定節日加班費",
        "multi_birth_count": "特殊計費: 胎數",
        "stair_floor_fee_mode": "透天服務樓層方式 (樓層加收費)",
        "parking_space_provided": "提供服務人員轎車停車位",
        "other_babies_present": "服務時間內是否有其他寶寶",
        "bank_code": "補助款退款:銀行代號+分行代號",
        "bank_account": "補助款退款:銀行帳號"
    },
    "global_stats (全域與多案件營運統計視圖)": {
        "global_active_orders_count": "📊 本週服務中案件總數",
        "global_active_staff_count": "📊 本週出勤月嫂總人數",
        "global_subsidy_orders_count": "📊 補助身份案件統計總數",
        "global_total_receivable_sum": "📊 系統應收自費總金額彙總",
        "global_govt_claim_count": "📊 本月市府請款案件總數"
    }
}

ALL_DB_FIELDS = {}
for _tbl, _fmap in DB_TABLE_FIELDS.items():
    ALL_DB_FIELDS.update(_fmap)


def get_table_for_key(db_key: str) -> str:
    """自動反查 db_key 隸屬哪一個 SQL 資料表」"""
    for tbl_name, fmap in DB_TABLE_FIELDS.items():
        if db_key in fmap:
            return tbl_name
    return list(DB_TABLE_FIELDS.keys())[0]


def render_html_document(tpl_data: dict, target_order: dict, global_stats: dict) -> str:
    """渲染雙欄 CSS Grid 高質感 HTML 實體單據"""
    html_items = ""
    for f in tpl_data.get('fields', []):
        lbl = f.get('label', '未命名欄位')
        f_type = f.get('type', 'text')
        f_width = f.get('width', 'half')
        val_str = "___________"

        if f_type == "db_link":
            db_k = f.get('db_key', 'client_name')
            if db_k in global_stats:
                val_raw = global_stats[db_k]
                val_str = format_db_value(db_k, val_raw)
            elif target_order:
                val_raw = target_order.get(db_k, '')
                val_str = format_db_value(db_k, val_raw)

        is_full = (f_width == "full" or f_type == "textarea")
        grid_span = "grid-column: span 2;" if is_full else ""

        html_items += f"""
        <div style="background:#FFFFFF; border:1px solid #E0E0E0; border-radius:6px; padding:10px 14px; {grid_span}">
            <div style="font-size:12px; color:#616161; font-weight:600; margin-bottom:4px;">• {lbl}</div>
            <div style="font-size:15px; color:#1565C0; font-weight:bold; word-break:break-word;">{val_str}</div>
        </div>
        """

    print_doc_html = f"""
    <div id="printable-area" style="font-family:'Segoe UI', Microsoft JhengHei, sans-serif; padding:25px; border:2px solid #1565C0; background:#FAFAFA; border-radius:8px; max-width:850px; margin:auto;">
        <div style="text-align:center; border-bottom:2px double #1565C0; padding-bottom:12px; margin-bottom:20px;">
            <h2 style="margin:0; color:#1565C0; font-size:22px;">新竹市月子照顧服務人員職業工會</h2>
            <h3 style="margin:6px 0 0 0; color:#333; font-size:17px;">【 {tpl_data.get('name', '表單單據')} 】</h3>
            <p style="margin:4px 0 0 0; color:#757575; font-size:11px;">單據編號：TPL-{tpl_data.get('id', '00')} | 列印日期：{date.today().strftime('%Y-%m-%d')}</p>
        </div>

        <div style="display:grid; grid-template-columns: 1fr 1fr; gap:10px 14px;">
            {html_items}
        </div>
    </div>

    <div style="text-align:center; margin-top:15px;">
        <button onclick="window.print()" style="background-color:#2E7D32; color:white; border:none; padding:10px 24px; font-size:15px; font-weight:bold; border-radius:6px; cursor:pointer; box-shadow:0 2px 4px rgba(0,0,0,0.2);">
            🖨️ 點擊另存為 PDF / 輸出圖片 (Print / Save as PDF)
        </button>
    </div>
    """
    return print_doc_html
